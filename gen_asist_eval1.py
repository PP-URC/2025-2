from faker import Faker
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import sys



fake = Faker('es_MX')
np.random.seed(42)
fake.seed_instance(42)



def generate_attendance(n_students, n_sessions, base_attendance=0.8, variability=0.15):

    student_rates = np.random.normal(base_attendance, variability, n_students)
    student_rates = np.clip(student_rates, 0.1, 0.98)
    attendance = np.zeros((n_students, n_sessions))

    for i, rate in enumerate(student_rates):
        attendance[i] = np.random.binomial(1, rate, n_sessions)

    return attendance.astype(int)

def generate_evals(n_students, n_assignments, min_score=5, max_score=10):
    """
    Generar evaluaciones con diferencias realistas entre estudiantes
    """
    scores_range = list(range(min_score, max_score + 1))
    
    # Crear diferentes perfiles de estudiantes
    scores = np.zeros((n_students, n_assignments))
    
    for i in range(n_students):
        # Asignar perfil de rendimiento basado en posición
        if i % 5 == 0:  # 20% - Rendimiento crítico
            base_prob = [0.25, 0.25, 0.2, 0.2, 0.08, 0.02]  # Sesgado hacia 5-7
            variability = 1.3  
        elif i % 5 == 1:  # 20% - Rendimiento bajo
            base_prob = [0.15, 0.2, 0.25, 0.2, 0.15, 0.05]  # Sesgado hacia 6-8
            variability = 1
        elif i % 5 == 2:  # 20% - Rendimiento inconsistente
            base_prob = [0.1, 0.15, 0.2, 0.2, 0.2, 0.15]  # Distribución plana
            variability = 2.0  # Muy alta variabilidad
        elif i % 5 == 3:  # 20% - Rendimiento promedio
            base_prob = [0.05, 0.1, 0.15, 0.25, 0.3, 0.15]  # Sesgado hacia 8-9
            variability = 0.8
        else:  # 20% - Rendimiento excelente
            base_prob = [0.02, 0.03, 0.05, 0.1, 0.3, 0.5]  # Sesgado hacia 9-10
            variability = 0.5  # Baja variabilidad
        
        varied_probs = np.array(base_prob) + np.random.normal(0, 0.7, 6)
        varied_probs = np.clip(varied_probs, 0.01, 0.99)
        varied_probs = varied_probs / varied_probs.sum()  
        
        student_scores = np.random.choice(scores_range, 
                                        size=n_assignments, 
                                        p=varied_probs)
        
        noise = np.random.normal(0, variability, n_assignments)
        student_scores = np.clip(student_scores + noise, min_score, max_score).astype(int)
        
        scores[i] = student_scores
    
    return scores




def data_to_excel_pd(directory, group, matriculas, names, evaluation_data, attendance_data):

    filename = f"{group}.xlsx"
    filepath = os.path.join(directory, filename)
    evaluation_df = pd.DataFrame(evaluation_data)
    evaluation_df.insert(0, 'Matricula', matriculas)
    evaluation_df.insert(1, 'Nombre', names)
    attendance_df = pd.DataFrame(attendance_data)
    attendance_df.insert(0, 'Matricula', matriculas)
    attendance_df.insert(1, 'Nombre', names)
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Save evaluations to second sheet
        evaluation_df.to_excel(writer, sheet_name='Evaluaciones', index=False)
        # Save attendance to first sheet
        attendance_df.to_excel(writer, sheet_name='Asistencia', index=False)



    print(f"✅  {len(names)} students to {filename}")

def data_to_excel_by_subject(directory, subject, group_data, n_lessons, n_evals):

    filename = f"{subject}.xlsx"
    filepath = os.path.join(directory, filename) 
    for group, data in group_data.items():
        evaluation_df = pd.DataFrame(data["evaluation_data"])        
        evaluation_df.insert(0, 'Matricula', data["matriculas"])
        evaluation_df.insert(1, 'Nombre', data["names"])
        attendance_df = pd.DataFrame(data["attendance_data"])
        attendance_df.insert(0, 'Matricula', data["matriculas"])
        attendance_df.insert(1, 'Nombre', data["names"])
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a' if os.path.exists(filepath) else 'w') as writer:
           
            

        
            evaluation_df.to_excel(writer, sheet_name=f'Evaluaciones_{group}', index=False)
      
            attendance_df.to_excel(writer, sheet_name=f'Asistencia_{group}', index=False)



    print(f"✅  {len(data["matriculas"])} students to {filename}")


def data_to_excel_by_subject(directory, subject, group_data, n_lessons, n_evals):
    filename = f"{subject}.xlsx"
    filepath = os.path.join(directory, filename)
    
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Crear workbook con TODAS las propiedades de Excel
    wb = Workbook()
    
    # ESTO ES CLAVE: Configurar propiedades del documento
    wb.properties.title = f"Datos Académicos - {subject}"
    wb.properties.subject = "Registro de Asistencia y Evaluaciones"
    wb.properties.creator = "Sistema Académico"
    wb.properties.keywords = "excel, educación, calificaciones"
    wb.properties.category = "Educación"
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    for group, data in group_data.items():
        # Evaluaciones
        eval_df = pd.DataFrame(data["evaluation_data"])        
        eval_df.insert(0, 'Matricula', data["matriculas"])
        eval_df.insert(1, 'Nombre', data["names"])
        
        ws_eval = wb.create_sheet(f'Evaluaciones_{group}')
        for row in dataframe_to_rows(eval_df, index=False, header=True):
            ws_eval.append(row)
        
        # Asistencia
        att_df = pd.DataFrame(data["attendance_data"])
        att_df.insert(0, 'Matricula', data["matriculas"])
        att_df.insert(1, 'Nombre', data["names"])
        
        ws_att = wb.create_sheet(f'Asistencia_{group}')
        for row in dataframe_to_rows(att_df, index=False, header=True):
            ws_att.append(row)
    
    # GUARDAR con todas las propiedades
    wb.save(filepath)
    
    # Verificar que se creó como Excel
    verify_excel_file(filepath)
    print(f"✅ {filename} creado como ARCHIVO EXCEL")

def verify_excel_file(filepath):
    """Verificar que el archivo es un Excel válido"""
    import magic  # pip install python-magic
    try:
        file_type = magic.from_file(filepath)
        print(f"   📄 Tipo detectado: {file_type}")
        return "Excel" in file_type or "Microsoft" in file_type
    except:
        # Si no tienes magic, verificar con pandas
        try:
            pd.read_excel(filepath)
            print(f"   ✅ Archivo Excel válido")
            return True
        except:
            print(f"   ❌ Archivo corrupto")
            return False



def generate_matriculas(num_students, start=264_421_500):
    matriculas = list(range(start, start + num_students))
    np.random.shuffle(matriculas)
    return matriculas

def generate_names(num_students):

    fake = Faker('es_MX')

    names, matriculas = [], []
    for _ in range(num_students):
        name = f"{fake.first_name()} {fake.last_name()} {fake.last_name()}"
        names.append(name)
    return names

def create_groups(n_groups, n_lessons=17, n_evals=10, min_students=10, max_students=25):
    
    groups = [f"Grupo{n}" for n in range(1, n_groups + 1)]
    populations = np.random.randint(min_students, max_students, 6)
    total_students = sum(populations)
    matriculas = generate_matriculas(total_students)
    # Create directory
    directory_name = 'asistencia_calificaciones'
    os.makedirs(directory_name, exist_ok=True)
    subjects = ["Calculo_Integral", "Bases_de_Datos", "Contabilidad_Financiera", "Estructuras_de_Datos", "Pensamiento_Complejo", "Probabilidad"]
    print("🎓 GENERANDO DATOS DE ESTUDIANTES POR GRUPO")

    for group, population in zip(groups, populations):
        count = 0
        names = generate_names(population)
        matriculas_group = matriculas[count:population]
        count += population
        for subject in subjects:
            attendance_data = generate_attendance(population, n_lessons)
            evaluation_data = generate_evals(population, n_evals)
            group_name = f"{subject}_{group}"
            data_to_excel_pd(directory_name, group_name, matriculas_group, names, evaluation_data, attendance_data)

def create_groups(n_groups=5, n_lessons=17, n_evals=10, min_students=10, max_students=25):
    """
    Structure:
    asistencia_calificaciones/
    ├── Calculo_Integral.xlsx
    │   ├── Grupo1_Calificaciones
    │   ├── Grupo1_Asistencia
    │   ├── Grupo2_Calificaciones
    │   └── ...
    ├── Bases_de_Datos.xlsx
    └── ...
    """
    groups = [f"Grupo{n}" for n in range(1, n_groups + 1)]
    populations = np.random.randint(min_students, max_students, n_groups)
    total_students = sum(populations)
    matriculas = generate_matriculas(total_students)
    
    # Create directory
    directory_name = 'asistencia_calificaciones'
    os.makedirs(directory_name, exist_ok=True)
    
    subjects = [
        "Calculo_Integral", 
        "Bases_de_Datos", 
        "Contabilidad_Financiera", 
        "Estructuras_de_Datos", 
        "Pensamiento_Complejo", 
        "Probabilidad"
    ]
    
    print("🎓 GENERANDO DATOS POR MATERIA")
    print(f"📚 Materias: {', '.join(subjects)}")
    print(f"👥 Grupos: {', '.join(groups)}")
    print(f"📊 Total de estudiantes: {total_students}")
    print("-" * 50)
    
    all_students_data = set()
    subject_data = {subject: {} for subject in subjects}
    
    count = 0
    for group, population in zip(groups, populations):
        names = generate_names(population)
        matriculas_group = matriculas[count:count + population]
        count += population
        for matricula, name in zip(matriculas_group, names):
            all_students_data.add((matricula, name, group))
        for subject in subjects:
            attendance_data = generate_attendance(population, n_lessons)
            evaluation_data = generate_evals(population, n_evals)
            
            subject_data[subject][group] = {
                'matriculas': matriculas_group,
                'names': names,
                'evaluation_data': evaluation_data,
                'attendance_data': attendance_data
            }
    
    for subject, group_data in subject_data.items():
        data_to_excel_by_subject(directory_name, subject, group_data, n_lessons, n_evals)
    
    print("-" * 50)
    print(f"📁 Archivos guardados en: {directory_name}/")
    return list(all_students_data)
